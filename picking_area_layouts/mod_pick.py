from __future__ import annotations

import pandas as pd
import re
from openpyxl import load_workbook

# ============================
# CONFIG (EDIT THESE)
# ============================

EXCEL_PATH = "../data/Heat Map - Pick Modules.xlsx"  # path to your Excel file
BIN_HEIGHT = 3                                  # bins are grouped in 3 cells (vertical OR horizontal)

# Only parse the real grid sheets (prevents tables like "Module Items", etc.)
# If you add Pick Module #3 later, it will still work.
SHEET_NAME_MUST_CONTAIN = "pick module"

# ============================
# REGEX DEFINITIONS
# ============================

# Location IDs look like M21-6-6C06
# Rule you gave:
#   - pick_module = FIRST digit after M  (e.g., M21 -> 2)
#   - floor       = number after first dash (e.g., M21-6-... -> 6)
LOCATION_RE = re.compile(r"M(\d)\d*-(\d+)-[A-Z0-9]+", re.I)

# ============================
# PART 1: EXCEL GRID -> LONG TABLE DF
# ============================

# Load workbook (data_only=True reads calculated values)
wb = load_workbook(EXCEL_PATH, data_only=True)

rows = []  # all parsed bin records will go here

print("Sheet names:", wb.sheetnames)

# Loop through every sheet (each may be a different pick module layout)
for sheet_name in wb.sheetnames:
    # Skip non-grid sheets
    if SHEET_NAME_MUST_CONTAIN not in sheet_name.lower():
        continue

    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    # ----------------------------
    # Find top-left bin anchor
    # ----------------------------
    start_row = None
    start_col = None

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and LOCATION_RE.fullmatch(val.strip()):
                start_row = r
                start_col = c
                break
        if start_row is not None:
            break

    # If no bins found in this sheet, skip it
    if start_row is None or start_col is None:
        continue

    # Walk column by column (x direction)
    # while-loop so we can advance col by BIN_HEIGHT when horizontal bins are found
    col = start_col
    while col <= max_col:
        y = 0  # vertical bin index per column
        row = start_row

        while row <= max_row:

            # ----------------------------
            # Check if this is a blank separator row
            # (floors separated by blank rows)
            # ----------------------------
            is_blank = True
            for i in range(BIN_HEIGHT):
                rr = row + i
                if rr > max_row:
                    continue
                val = ws.cell(row=rr, column=col).value
                if val not in (None, "", "OPEN"):
                    is_blank = False
                    break

            # Blank → move down 1 row and re-align
            if is_blank:
                row += 1
                continue

            # ============================
            # Try VERTICAL stack first
            # ============================
            stack = []
            for i in range(BIN_HEIGHT):
                rr = row + i
                if rr > max_row:
                    stack.append(None)
                    continue

                cell = ws.cell(row=rr, column=col)
                value = cell.value

                # merged-cell safe
                if value is None:
                    for merged in ws.merged_cells.ranges:
                        if (
                            merged.min_row <= rr <= merged.max_row
                            and merged.min_col <= col <= merged.max_col
                        ):
                            value = ws.cell(row=merged.min_row, column=merged.min_col).value
                            break

                stack.append(value)

            # ============================
            # Classify stack (order doesn't matter)
            # Handles numeric SKU + numeric quantity:
            #  - if 2+ numerics: smallest = quantity, largest = sku (if sku not already a string)
            # ============================
            location_id = None
            sku = None
            quantity = None

            num_vals = []
            str_vals = []

            for v in stack:
                if isinstance(v, str):
                    vv = v.strip()
                    if vv == "" or vv.upper() == "OPEN":
                        continue
                    if LOCATION_RE.fullmatch(vv):
                        location_id = vv
                    else:
                        str_vals.append(vv)
                elif isinstance(v, (int, float)):
                    num_vals.append(float(v))
                elif isinstance(v, str) and v.strip().isdigit():
                    num_vals.append(float(v.strip()))

            if str_vals:
                sku = str_vals[0]

            if len(num_vals) == 1:
                quantity = num_vals[0]
            elif len(num_vals) >= 2:
                num_vals_sorted = sorted(num_vals)
                quantity = num_vals_sorted[0]
                if sku is None:
                    sku = str(int(num_vals_sorted[-1]))

            horizontal = False  # assume vertical unless horizontal succeeds

            # ============================
            # If vertical failed → try HORIZONTAL stack
            # ============================
            if location_id is None or quantity is None:
                stack = []
                for i in range(BIN_HEIGHT):
                    cc = col + i
                    if cc > max_col:
                        stack.append(None)
                        continue

                    cell = ws.cell(row=row, column=cc)
                    value = cell.value

                    # merged-cell safe
                    if value is None:
                        for merged in ws.merged_cells.ranges:
                            if (
                                merged.min_row <= row <= merged.max_row
                                and merged.min_col <= cc <= merged.max_col
                            ):
                                value = ws.cell(row=merged.min_row, column=merged.min_col).value
                                break

                    stack.append(value)

                # Re-classify horizontal stack
                location_id = None
                sku = None
                quantity = None

                num_vals = []
                str_vals = []

                for v in stack:
                    if isinstance(v, str):
                        vv = v.strip()
                        if vv == "" or vv.upper() == "OPEN":
                            continue
                        if LOCATION_RE.fullmatch(vv):
                            location_id = vv
                        else:
                            str_vals.append(vv)
                    elif isinstance(v, (int, float)):
                        num_vals.append(float(v))
                    elif isinstance(v, str) and v.strip().isdigit():
                        num_vals.append(float(v.strip()))

                if str_vals:
                    sku = str_vals[0]

                if len(num_vals) == 1:
                    quantity = num_vals[0]
                elif len(num_vals) >= 2:
                    num_vals_sorted = sorted(num_vals)
                    quantity = num_vals_sorted[0]
                    if sku is None:
                        sku = str(int(num_vals_sorted[-1]))

                horizontal = (location_id is not None and quantity is not None)

            # Not a valid bin → slide down by 1 row
            if location_id is None or quantity is None:
                row += 1
                continue

            # ----------------------------
            # Parse module + floor
            # ----------------------------
            m = LOCATION_RE.fullmatch(location_id)
            pick_module = int(m.group(1))
            floor = int(m.group(2))

            rows.append({
                "location_id": location_id,
                "sku": sku,
                "quantity": quantity,
                "x": col - start_col,   # x relative to anchor
                "y": y,
                "pick_module": pick_module,
                "floor": floor,
                "sheet": sheet_name,
            })

            y += 1

            # Advance pointer
            if horizontal:
                col += BIN_HEIGHT
            else:
                row += BIN_HEIGHT

        # If we finished scanning down for this column, move to next column
        col += 1


# Output of Part 1 (this is the INPUT to Part 2)
df_part1 = pd.DataFrame(rows)
print("\nPart 1 output (head):")
print(df_part1.head())

# Optional save
df_part1.to_csv("warehouse_bins_long.csv", index=False)

# ============================
# PART 2: PART 1 DF -> CLEAN / VALIDATED DF FOR USE
# (This replaces the old bins.csv schema with your new schema)
# ============================

REQUIRED_COLS = ["pick_module", "floor", "x", "y", "location_id", "sku", "quantity"]


def bins_df_to_clean_indexed_df(df: pd.DataFrame, *, validate: bool = True) -> pd.DataFrame:
    """
    Take Part 1 output df and return a clean DataFrame with:
      - required columns present (order doesn't matter)
      - numeric types for pick_module/floor/x/y/quantity
      - optional validation
      - indexed by (pick_module, floor, x, y) for fast lookup
    """
    # Column check
    missing = set(REQUIRED_COLS) - set(df.columns)
    extra = set(df.columns) - set(REQUIRED_COLS)  # you can relax this if you want to keep 'sheet'
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    # Keep required columns (drop extras like 'sheet' unless you want to keep them)
    df = df[REQUIRED_COLS].copy()

    # Normalize whitespace in string columns
    df["location_id"] = df["location_id"].astype(str).str.strip()
    df["sku"] = df["sku"].astype(str).str.strip()
    df.loc[df["sku"].isin(["", "nan", "None"]), "sku"] = None  # normalize empties

    # Convert numeric columns
    for c in ["pick_module", "floor", "x", "y"]:
        df[c] = pd.to_numeric(df[c], errors="raise").astype(int)

    # quantity: keep as float (can be int-like, but safer as float)
    df["quantity"] = pd.to_numeric(df["quantity"], errors="raise").astype(float)

    if validate:
        if (df["location_id"] == "").any():
            bad = df.index[df["location_id"] == ""].tolist()[:10]
            raise ValueError(f"Blank location_id found in rows: {bad} (showing up to 10)")

        if (df["quantity"] < 0).any():
            raise ValueError("quantity cannot be negative")

        # Uniqueness: usually location_id should be unique in a grid
        if df["location_id"].duplicated().any():
            dups = df.loc[df["location_id"].duplicated(), "location_id"].unique()[:10]
            raise ValueError(f"Duplicate location_id(s): {dups}")

        # Coordinate uniqueness (per module/floor)
        coord_cols = ["pick_module", "floor", "x", "y"]
        if df.duplicated(subset=coord_cols).any():
            dups = df.loc[df.duplicated(subset=coord_cols), coord_cols].head(10)
            raise ValueError(f"Duplicate (pick_module,floor,x,y) rows detected:\n{dups}")

    # Set index for fast lookup in your simulation
    df = df.set_index(["pick_module", "floor", "x", "y"]).sort_index()
    return df
